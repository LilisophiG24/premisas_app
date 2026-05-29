import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
import copy

st.markdown("""
<style>
.week-badge {background:#1F3864;color:white;padding:6px 18px;border-radius:6px;
             font-size:1.2rem;font-weight:bold;display:inline-block;}
.section-title {color:#1F3864;font-weight:bold;font-size:1rem;margin-bottom:4px;}
.tag-r  {background:#d4edda;color:#155724;padding:1px 6px;border-radius:3px;font-size:.8rem;font-weight:bold;}
.tag-i  {background:#fff3cd;color:#856404;padding:1px 6px;border-radius:3px;font-size:.8rem;font-weight:bold;}
.tag-ri {background:#cce5ff;color:#004085;padding:1px 6px;border-radius:3px;font-size:.8rem;font-weight:bold;}

/* Data editor: better contrast and auto row height */
div[data-testid="stDataFrame"] .ag-cell {
    color: #ffffff !important;
    font-weight: 500 !important;
    white-space: normal !important;
    line-height: 1.4 !important;
    padding-top: 6px !important;
    padding-bottom: 6px !important;
}
div[data-testid="stDataFrame"] .ag-cell-value {
    white-space: normal !important;
    overflow: visible !important;
    word-break: break-word !important;
}
div[data-testid="stDataFrame"] .ag-row {
    height: auto !important;
    min-height: 42px !important;
}
div[data-testid="stDataFrame"] .ag-header-cell-label {
    font-weight: 700 !important;
    color: #ffffff !important;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════
DIAS_SEMANA = ["Sábado","Domingo","Lunes","Martes","Miércoles","Jueves","Viernes"]

# Equipment types that indicate INDISPONIBILIDAD (generation)
I_EQUIP_TYPES = [
    "GENERADOR", "UNIDAD DE GENERACION", "GRUPO GENERADOR",
    "INVERSOR ESTATICO", "SEDIMENTADOR", "PRESA", "CANAL DE TRASVASE",
]
# If PLANTA prefix + TRANSFORMADOR → also I
I_DESC_KW = [
    "mantenimiento", "falla", "reparacion", "reparación", "reemplazo",
    "cambio de rodete", "cambio", "inspeccion", "inspección",
    "revision", "revisión", "averia", "avería", "paro", "parada",
    "correctivo", "overhaul", "desmontaje", "trabajos"
]

# Equipment types that indicate RELEVANTE (transmission)
R_EQUIP_TYPES = [
    "LINEA DE CONEXION", "AUTOTRANSFORMADOR", "BANCO DE CAPACITORES",
    "BARRA", "PORTICO", "PÓRTICO", "TRANSFORMADOR DE POTENCIA"
]
R_VOLT_RE  = re.compile(r'\b(230|115|34\.5|34)-|\b(230|115|34\.5)\s*k[Vv]', re.I)
R_LINE_RE  = re.compile(r'\bLIN\s+(230|115|34)[-\s]', re.I)
R_DESC_KW  = [
    "reemplazo", "cambio de aisladores", "tendido", "opgw", "conductor",
    "prueba", "pruebas electricas", "pruebas eléctricas",
    "mantenimiento", "portico", "pórtico", "torre", "aislador",
    "herraje", "poda", "reemplazar", "interruptor"
]

# Plantilla column maps
PLANTILLA_LIB_COLS = [
    "Número","Tipo","Es repetitiva","Fecha Solicitud","Fecha Inicio",
    "Fecha Final","Duración Programada","Fecha Inicio Real","Fecha Final Real",
    "Duración Real","Descripción","Observaciones","Equipos",
    "Libranzas Vinculadas","Fecha Aprobación","Último Estado","R/I"
]
PLANTILLA_REL_COLS = [
    "Tipo","Fecha inicio","Hora inicio","Fecha final","Hora final",
    "Tipo de Equipos","Equipo","Subestación","Libranza",
    "Descripción del trabajo","Estado","Observaciones"
]
INDISP_COLS = [
    "Fecha inicio","Hora inicio","Fecha final","Hora final",
    "Unidad","Potencia (MW)","Libranza","Descripción"
]
PROY_COLS = [
    "Planta","Tecnología","MW","Fecha Planeamiento",
    "Sem Disp","Sem Prueba","Última Solicitud","Fecha finaliza libranza"
]

# ══════════════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════════════
def parse_dt(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    if isinstance(val, date): return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    for fmt in ["%d/%m/%Y %H:%M","%d/%m/%Y","%Y-%m-%d %H:%M","%Y-%m-%d"]:
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

def fmt_date(dt):
    if dt is None: return ""
    return dt.strftime("%d/%m/%Y")

def fmt_time(dt):
    if dt is None: return ""
    return dt.strftime("%H:%M")

def week_range(dates):
    valid = [d for d in dates if d]
    if not valid: return "—"
    s,e = valid[0], valid[-1]
    meses = ["","ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]
    return f"{s.day} {meses[s.month]} – {e.day} {meses[e.month]} {e.year}"

# ══════════════════════════════════════════════════════════════════════
# CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════
SCADA_KW = [
    "scada", "rtu", "comunicaciones", "fibra", "telecomunicaciones",
    "fibra óptica", "fibra optica", "red de comunicacion", "red de comunicación",
    "enlace", "radiocomunicacion", "radiocomunicación", "microondas",
]

DISTRIBUTION_AGENTS = ["ENSA-", "EDEMET-", "NATURGY-", "EDECHI-"]
DISTRIBUTION_KW = ["circuito"]

SPECIAL_PLANTS = {
    'FORTUNA','BAYANO','COSTA NORTE','CRISTOBAL',
    'ESTI','ESTÍ','GATÚN','GATUN','GENERADORA GATUN','GATUN II'
}

@st.cache_data(show_spinner=False)
def load_equipos_db(file_bytes):
    """Load equipos_libranzas.xlsx → (db, etesa_codes, etesa_locs)."""
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        db, etesa_codes, etesa_locs = {}, set(), set()
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]: continue
            code   = str(row[1] or "").strip().upper()
            agent  = str(row[2] or "").strip().upper()
            instal = str(row[3] or "").strip().upper()
            tipo   = str(row[4] or "").strip().upper()
            extras = str(row[5] or "").strip().upper()
            if not code: continue
            db.setdefault(code, []).append((agent, instal, tipo, extras))
            if agent == "ETESA" or "ETESA" in extras:
                etesa_codes.add(code)
                if "->" in instal:
                    loc = instal.split("->")[-1].strip()
                    if loc: etesa_locs.add(loc)
        return db, etesa_codes, etesa_locs
    except: return {}, set(), set()

def build_lineas_set(ws_lineas):
    """Build set of equipment codes from Lineas sheet (col B)."""
    codes = set()
    for row in ws_lineas.iter_rows(min_row=1, values_only=True):
        if row[1] and str(row[1]).strip() not in ('-',''):
            for code in str(row[1]).split('\n'):
                code = code.strip()
                if code: codes.add(code.upper())
    return codes

def parse_unit_mw_lookup(ws_indisp):
    """Parse unit code → MW reference table from Indisponibilidades sheet.
    Returns: (unit_mw dict, plant_prefix dict)."""
    unit_mw = {}      # {'BAYG3': 67.55, ...}
    plant_prefix = {} # {'bayano': 'BAY', 'fortuna': 'FOR', ...}
    current_prefix = None
    current_units  = []

    for row in ws_indisp.iter_rows(min_row=100, max_row=ws_indisp.max_row, max_col=8, values_only=True):
        if row[5] is None: continue
        code = str(row[5]).strip()
        mw   = float(row[6] or 0)

        # Unit code pattern: 2–5 uppercase letters + G + digits
        if re.match(r'^[A-Z]{2,5}G\d+$', code):
            unit_mw[code] = mw
            current_units.append(code)
            m = re.match(r'^([A-Z]+)G\d+$', code)
            if m: current_prefix = m.group(1)
        else:
            # Total / plant name row
            plant_name = re.sub(r'^Total\s+', '', code).strip().lower()
            if current_prefix and plant_name:
                plant_prefix[plant_name] = current_prefix
            unit_mw[code] = mw
            current_units  = []
            current_prefix = None

    return unit_mw, plant_prefix

def get_potencia_for_libranza(row, unit_mw, plant_prefix):
    """Sum MW for all generation units in the libranza's Equipos field."""
    equipos = str(row.get("Equipos","") or "")
    total_mw = 0.0
    for line in equipos.split('\n'):
        line = line.strip()
        m = re.match(r'^PLANTA\s+(.+?)\s*->\s*(.+)', line, re.I)
        if not m: continue
        plant_name = m.group(1).strip().lower()
        rest = m.group(2)
        prefix = next(
            (v for k, v in plant_prefix.items()
             if k in plant_name or plant_name in k),
            None
        )
        if not prefix: continue
        for uid in re.findall(r'G\d+', rest, re.I):
            code = prefix + uid.upper()
            total_mw += unit_mw.get(code, 0.0)
    return total_mw

def extract_equip_codes(equipos_str):
    """Extract all equipment codes from Equipos field for Lineas set lookup."""
    codes = set()
    for line in str(equipos_str or "").split('\n'):
        line = line.strip()
        if '->' not in line: continue
        # Left side: location (may have line codes like "LIN 230-5A")
        loc = line.split('->')[0].strip()
        for tok in re.findall(r'\d+[-\w\.]+', loc):
            codes.add(tok.upper())
        # Right side: equipment codes after ':'
        right = line.split('->')[1]
        for part in right.split(':'):
            for tok in re.findall(r'[\w][-\w\.]*', part):
                if re.search(r'[A-Z]', tok): codes.add(tok.upper())
    return codes

def get_unit_name_and_mw(row, unit_mw, plant_prefix):
    """Return (unit_label, total_mw) for an I libranza."""
    equipos = str(row.get("Equipos","") or "")
    labels = []
    total_mw = 0.0
    for line in equipos.split('\n'):
        line = line.strip()
        m = re.match(r'^PLANTA\s+(.+?)\s*->\s*(.+)', line, re.I)
        if not m: continue
        plant_name = m.group(1).strip()
        rest = m.group(2)
        plant_lower = plant_name.lower()
        prefix = next(
            (v for k, v in plant_prefix.items()
             if k in plant_lower or plant_lower in k),
            None
        )
        unit_ids = re.findall(r'G\d+', rest, re.I)
        for uid in unit_ids:
            code = (prefix or plant_name[:3].upper()) + uid.upper()
            mw = unit_mw.get(code, 0.0)
            total_mw += mw
            labels.append(code)
    return (' '.join(labels) if labels else get_unit_from_equipos(equipos)), total_mw

def classify_libranza(row, lineas_codes=None, unit_mw=None, plant_prefix=None, equipos_db=None):
    """Return 'R', 'I', 'R-I' or '' for a libranza row."""
    numero = str(row.get("Número","") or "").upper()
    eq     = str(row.get("Equipos","") or "")
    eq_up  = eq.upper()
    desc   = str(row.get("Descripción","") or "").lower()
    is_i   = False
    is_r   = False

    # ── Hard excludes — never R or I ───────────────────────────
    if "libranza informativa" in desc or "lib. informativa" in desc:
        return ""
    is_distrib = any(numero.startswith(a) for a in DISTRIBUTION_AGENTS)
    if is_distrib and any(k in desc or k in eq_up.lower() for k in DISTRIBUTION_KW):
        return ""
    # SCADA/fibra/comunicaciones → never R regardless of agent or equipos_db
    if any(k in desc for k in SCADA_KW):
        return ""

    # ── I: non-ETESA ───────────────────────────────────────────
    if not numero.startswith("ETESA-"):
        is_i = True

    # ── R: ETESA equipment checked against Lineas sheet ────────
    if numero.startswith("ETESA-"):
        # GENERADOR AUXILIAR alone → never R for ETESA
        eq_codes_etesa = extract_equip_codes(eq)
        eq_codes_etesa -= {"GENERADOR AUXILIAR", "GEN AUX", "GENERADOR AUX"}
        only_genaux = not eq_codes_etesa and "GENERADOR AUXILIAR" in eq_up
        if not only_genaux:
            if lineas_codes and eq_codes_etesa & lineas_codes:
                is_r = True
            if not is_r:
                if R_LINE_RE.search(eq_up): is_r = True
            if not is_r and "BANCO DE CAPACITORES" in eq_up: is_r = True
            # STATCOM/SPEAR only if NOT inside a GENERADOR AUXILIAR name
            if not is_r:
                if ("STATCOM" in eq_up and "GENAUX" not in eq_up
                        and "GENERADOR AUXILIAR" not in eq_up): is_r = True
                if "SPEAR" in eq_up: is_r = True
            if not is_r:
                for t in R_EQUIP_TYPES:
                    if t in eq_up and R_VOLT_RE.search(eq_up):
                        is_r = True; break

    # ── R: Special large plants → always R+I ───────────────────
    if is_i and not is_r:
        for line in eq.split('\n'):
            m = re.match(r'^PLANTA\s+(.+?)\s*->', line.strip(), re.I)
            if m:
                pname = m.group(1).strip().upper()
                if any(sp in pname or pname in sp for sp in SPECIAL_PLANTS):
                    is_r = True; break

    # ── R: equipos_db — equipment owned/shared by ETESA ────────
    # Only applies if not GENERADOR AUXILIAR only
    if not is_r and equipos_db:
        _, etesa_codes_db, etesa_locs_db = equipos_db
        eq_codes_db = extract_equip_codes(eq)
        # Remove GENERADOR AUXILIAR codes from check
        eq_codes_db -= {"GENERADOR AUXILIAR", "GEN AUX", "GENERADOR AUX"}
        if eq_codes_db and eq_codes_db & etesa_codes_db:
            is_r = True
        if not is_r and eq_codes_db:
            for loc in etesa_locs_db:
                if loc and len(loc) > 3 and loc in eq_up:
                    is_r = True; break

    # ── R: Other agents >60 MW ─────────────────────────────────
    if is_i and not is_r and unit_mw and plant_prefix:
        if get_potencia_for_libranza(row, unit_mw, plant_prefix) > 60:
            is_r = True

    if is_r and is_i: return "R-I"
    if is_r: return "R"
    if is_i: return "I"
    return ""

# ══════════════════════════════════════════════════════════════════════
# EQUIPOS PARSER
# ══════════════════════════════════════════════════════════════════════
TIPO_MAP = {
    "LINEA DE CONEXION":"Línea","AUTOTRANSFORMADOR":"Transformador",
    "TRANSFORMADOR DE POTENCIA":"Transformador",
    "TRANSFORMADOR PUESTO A TIERRA":"Transformador","TRANSFORMADOR":"Transformador",
    "BANCO DE CAPACITORES":"Banco de Capacitores",
    "BANCO DE BATERÍAS":"Banco de Baterías","BANCO DE BATERIAS":"Banco de Baterías",
    "BARRA":"Barra","PORTICO":"Pórtico","PÓRTICO":"Pórtico",
    "INTERRUPTOR":"Interruptor","GENERADOR AUXILIAR":"Generador Auxiliar",
    "GENERADOR":"Generador","UNIDAD DE GENERACION":"Unidad de Generación",
    "GRUPO GENERADOR":"Grupo Generador","INVERSOR ESTATICO":"Inversor",
    "CUCHILLA MANUAL DE TIERRA":"Cuchilla","CUCHILLA MOTORIZADA":"Cuchilla",
    "CUCHILLA DE ATERRIZAJE":"Cuchilla","CUCHILLA MANUAL":"Cuchilla",
    "PRESA":"Presa","SEDIMENTADOR":"Sedimentador","CIRCUITO":"Circuito",
    "RELEVADOR":"Relevador","MEDIDOR":"Medidor","TABLERO DE CONTROL":"Tablero",
    "COMPUTADOR":"Computador","RTU":"RTU","SECCIONADOR":"Seccionador",
}

_ALL_TYPES_SORTED = sorted(TIPO_MAP.keys(), key=len, reverse=True)
_TYPE_RE = re.compile(
    r'(' + '|'.join(re.escape(t) for t in _ALL_TYPES_SORTED) + r'):\s*',
    re.IGNORECASE
)
_LOC_PREFIX_RE = re.compile(r'\s{2,}(?=(?:SE|LIN|PLANTA)\s+\S.*?->)')

def parse_equipo_entries(equipos_str):
    """Parse Equipos string into list of {location, raw_tipo, tipo, ids}."""
    results = []
    if not equipos_str: return results
    # Normalize: multiple entries on one line separated by 3+ spaces
    normalized = _LOC_PREFIX_RE.sub('\n', str(equipos_str))
    for line in normalized.split("\n"):
        line = line.strip()
        if not line or "->" not in line: continue
        m = re.match(r'^(.+?)\s*->\s*(.+)$', line)
        if not m: continue
        location = m.group(1).strip()
        rest = m.group(2).strip()
        matches = list(_TYPE_RE.finditer(rest))
        for i, tm in enumerate(matches):
            raw_tipo = tm.group(1).upper()
            start = tm.end()
            end = matches[i+1].start() if i+1 < len(matches) else len(rest)
            ids = rest[start:end].strip()
            tipo_friendly = TIPO_MAP.get(raw_tipo, raw_tipo.title())
            results.append({"location": location, "raw_tipo": raw_tipo,
                            "tipo": tipo_friendly, "ids": ids})
    return results

def get_relevante_equipo_info(row, lineas_lookup):
    """Extract Tipo de Equipos, ALL Equipos, Subestación for relevantes table."""
    entries = parse_equipo_entries(row.get("Equipos",""))
    R_PRIORITY = ["LINEA DE CONEXION","AUTOTRANSFORMADOR","TRANSFORMADOR DE POTENCIA",
                  "BANCO DE CAPACITORES","BARRA","PORTICO","PÓRTICO","INTERRUPTOR"]

    # Collect ALL R-type entries
    r_entries = []
    for priority in R_PRIORITY:
        for e in entries:
            if priority in e["raw_tipo"] and e not in r_entries:
                r_entries.append(e)
    if not r_entries and entries:
        r_entries = entries[:1]
    if not r_entries:
        return {"tipo":"—","equipo":"—","sust":"—"}

    # Tipo: from first R entry
    tipo = r_entries[0]["tipo"]

    # All equipo IDs combined
    all_ids = "  ".join(
        re.sub(r'\(RTR\)', '', e["ids"]).strip()
        for e in r_entries
    ).strip()

    # Subestación: from first line entry lookup, else location
    sust = ""
    for e in r_entries:
        if "LINEA DE CONEXION" in e["raw_tipo"]:
            for token in re.sub(r'\(RTR\)', '', e["ids"]).split():
                if token.strip() in lineas_lookup:
                    sust = lineas_lookup[token.strip()]
                    break
        if sust: break
    if not sust:
        loc = r_entries[0]["location"]
        loc = re.sub(r'^(SE|LIN|PLANTA)\s+', '', loc, flags=re.I).strip()
        sust = loc.title()

    return {"tipo": tipo, "equipo": all_ids, "sust": sust}

def get_unit_from_equipos(equipos_str):
    """Extract generation unit description for indisponibilidades."""
    entries = parse_equipo_entries(equipos_str)
    GEN_TYPES = ["GENERADOR","UNIDAD DE GENERACION","GRUPO GENERADOR","INVERSOR ESTATICO"]
    for e in entries:
        if any(t in e["raw_tipo"] for t in GEN_TYPES):
            loc = re.sub(r'^PLANTA\s+', '', e["location"], flags=re.I).strip()
            return f"{e['ids']}  ({loc})"
    if entries:
        e = entries[0]
        return f"{e['ids']}  ({e['location']})"
    return equipos_str[:60] if equipos_str else ""

# ══════════════════════════════════════════════════════════════════════
# LOADERS
# ══════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def load_plantilla(file_bytes):
    """Load template workbook and extract calendar, line lookup, proyectos."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    out = {"wb_bytes": file_bytes}

    # ── Week calendar from Datos ───────────────────────────────────
    ws = wb["Datos"]
    weeks = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=8, values_only=True):
        label = str(row[0] or "").strip()
        if label.startswith("Semana"):
            try:
                num = int(label.split()[1])
                if num in weeks:
                    continue  # keep first occurrence (current year), skip duplicates
                dates = [parse_dt(row[i]) for i in range(1,8)]
                weeks[num] = dates  # [Sab,Dom,Lun,Mar,Mie,Jue,Vie]
            except: pass
    out["weeks"] = weeks

    # ── Current week = plantilla week + 1 ────────────────────────────
    ws_i = wb["Indisponibilidades"]
    title = str(ws_i.cell(1,1).value or "")
    m = re.search(r"Semana\s+(\d+)", title)
    plantilla_week = int(m.group(1)) if m else 1
    out["current_week"] = plantilla_week + 1  # generate NEXT week

    # ── Lineas lookup: code → description ─────────────────────────
    lineas = {}
    ws_l = wb["Lineas"]
    for row in ws_l.iter_rows(min_row=1, values_only=True):
        if row[1] and row[2]:
            for code in str(row[1]).split("\n"):
                code = code.strip()
                if code:
                    lineas[code] = str(row[2]).strip()
    out["lineas_lookup"] = lineas

    # ── Proyectos de Generación ────────────────────────────────────
    ws_p = wb["Proyectos de Generacion"]
    proy_rows = []
    hdr_found = False
    for row in ws_p.iter_rows(min_row=13, values_only=True):
        if not hdr_found:
            if row[0] == "Planta": hdr_found = True
            continue
        if not any(c is not None for c in row[:8]): continue
        r = list(row) + [None]*9
        fecha_plan = r[3]
        fecha_fin  = r[7]
        proy_rows.append({
            "Planta":     str(r[0] or ""),
            "Tecnología": str(r[1] or ""),
            "MW":         r[2],
            "Fecha Planeamiento": fmt_date(parse_dt(fecha_plan)) if fecha_plan else str(r[3] or ""),
            "Sem Disp":   str(r[4] or ""),
            "Sem Prueba": str(r[5] or ""),
            "Última Solicitud": str(r[6] or ""),
            "Fecha finaliza libranza": fmt_date(parse_dt(fecha_fin)) if isinstance(fecha_fin, datetime) else str(r[7] or ""),
            "_fecha_fin_raw": fecha_fin,
        })
    out["proyectos"] = pd.DataFrame(proy_rows) if proy_rows else pd.DataFrame(columns=PROY_COLS+["_fecha_fin_raw"])

    # ── Unit MW reference table ────────────────────────────────────
    unit_mw, plant_prefix = parse_unit_mw_lookup(wb["Indisponibilidades"])
    out["unit_mw"]       = unit_mw
    out["plant_prefix"]  = plant_prefix

    # ── Lineas set for R classification ──────────────────────────
    out["lineas_codes"] = build_lineas_set(wb["Lineas"])

    # ── Existing Indisponibilidades (flat unique list) ─────────────
    out["indisp_existing"] = parse_existing_indisp_flat(wb["Indisponibilidades"])

    # ── Previous week's Libranzas Relevantes (carry forward) ──────
    out["relevantes_anteriores"] = parse_relevantes_anteriores(wb["Libranzas Relevantes"])

    return out

def parse_relevantes_anteriores(ws):
    """Parse previous week's Libranzas Relevantes sheet → list of dicts."""
    rows = []
    hdr_found = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not hdr_found:
            if row[0] == "Tipo": hdr_found = True
            continue
        if not any(c is not None for c in row[:9]): continue
        r = list(row) + [None]*12
        rows.append({
            "Tipo":               str(r[0] or ""),
            "Fecha inicio":       str(r[1] or ""),
            "Hora inicio":        str(r[2] or ""),
            "Fecha final":        str(r[3] or ""),
            "Hora final":         str(r[4] or ""),
            "Tipo de Equipos":    str(r[5] or ""),
            "Equipo":             str(r[6] or ""),
            "Subestación":        str(r[7] or ""),
            "Libranza":           str(r[8] or ""),
            "Descripción del trabajo": str(r[9] or ""),
            "Estado":             str(r[10] or ""),
            "Observaciones":      str(r[11] or ""),
            "_sort_dt":           parse_dt(str(r[1] or "")) or datetime.max,
        })
    return rows

def parse_existing_indisp_flat(ws):
    """Parse plantilla's Indisponibilidades → flat unique list of entries (one per libranza)."""
    seen = set()
    entries = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(c is not None for c in row[:9]): continue
        if row[1] and "INDISPONIBILIDAD TOTAL" in str(row[1]): continue
        if isinstance(row[0], datetime): continue  # day header
        if row[7] is None: continue
        lib = str(row[7]).strip()
        if not lib or lib in seen: continue
        seen.add(lib)
        entries.append({
            "Fecha inicio":  str(row[1] or ""),
            "Hora inicio":   str(row[2] or ""),
            "Fecha final":   str(row[3] or ""),
            "Hora final":    str(row[4] or ""),
            "Unidad":        str(row[5] or ""),
            "Potencia (MW)": float(row[6] or 0),
            "Libranza":      lib,
            "Descripción":   str(row[8] or ""),
        })
    return entries

def parse_existing_indisp(ws):  # kept for backward compat
    return parse_existing_indisp_flat(ws)

@st.cache_data(show_spinner=False)
def load_source_libranzas(file_bytes):
    """Load libranzas source file using header names, not positional indices."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    headers = {}
    hdr_found = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not hdr_found:
            if row[0] == "Número":
                hdr_found = True
                headers = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
            continue
        if not any(c is not None for c in row): continue

        def g(col, default=""):
            idx = headers.get(col)
            if idx is None: return default
            val = row[idx] if idx < len(row) else None
            return str(val) if val is not None else default

        rows.append({
            "Número":            row[0],
            "Tipo":              g("Tipo"),
            "Es repetitiva":     g("Es repetitiva"),
            "Agente":            g("Agente"),
            "Fecha Solicitud":   g("Fecha Solicitud"),
            "Solicitante":       g("Solicitante"),
            "Fecha Inicio":      g("Fecha Inicio"),
            "Fecha Final":       g("Fecha Final"),
            "Duración Programada": g("Duración Programada"),
            "Fecha Inicio Real": g("Fecha Inicio Real"),
            "Fecha Final Real":  g("Fecha Final Real"),
            "Duración Real":     g("Duración Real"),
            "Descripción":       g("Descripción"),
            "Observaciones":     g("Observaciones"),
            "Equipos":           g("Equipos"),
            "Libranzas Vinculadas": g("Libranzas Vinculadas"),
            "Responsable de Campo": g("Responsable de Campo"),
            "Cargo Responsable Campo": g("Cargo Responsable Campo"),
            "Fecha Aprobación":  g("Fecha Aprobación"),
            "Último Estado":     g("Último Estado"),
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame()

@st.cache_data(show_spinner=False)
def load_indisp_file(file_bytes):
    """Parse INDISPONIBILIDAD DE UNIDADES file.
    Cols: B=Unidad, C=Fecha salida, D=Causa, E=Libranza, F=Fecha entrada, G=Potencia(MW)
    Returns flat DataFrame with INDISP_COLS + ['status']."""
    MESES_ES = {'ene':'01','feb':'02','mar':'03','abr':'04','may':'05','jun':'06',
                'jul':'07','ago':'08','sep':'09','oct':'10','nov':'11','dic':'12'}

    def parse_es_date(val):
        if val is None: return None
        s = str(val).strip()
        # Try standard formats first
        for fmt in ["%Y-%m-%d %H:%M:%S","%Y-%m-%d %H:%M","%Y-%m-%d"]:
            try: return datetime.strptime(s[:len(fmt)+2].strip(), fmt)
            except: pass
        # Spanish: "26-mar-2026 09:30"
        m = re.match(r'(\d{1,2})-([a-záéíóú]+)-(\d{4})(?:\s+(\d{1,2}:\d{2}))?', s, re.I)
        if m:
            mes = MESES_ES.get(m.group(2).lower()[:3], '01')
            hora = m.group(4) or '00:00'
            try: return datetime.strptime(f"{m.group(1):0>2}/{mes}/{m.group(3)} {hora}", "%d/%m/%Y %H:%M")
            except: pass
        return None

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = []
        hdr_found = False
        for row in ws.iter_rows(min_row=1, values_only=True):
            # Find header row (col B = "UNIDAD")
            if not hdr_found:
                if str(row[1] or '').strip().upper() == 'UNIDAD':
                    hdr_found = True
                continue
            b = row[1]; c = row[2]; d = row[3]; e = row[4]; f = row[5]; g = row[6]
            if b is None or str(b).strip() in ('','▼','▲'): continue
            if 'TOTAL' in str(b or '').upper() or 'ZONA' in str(b or '').upper(): continue
            fi = parse_es_date(c)
            ff = parse_es_date(f)
            if fi is None: continue
            rows.append({
                "Fecha inicio":  fmt_date(fi),
                "Hora inicio":   fmt_time(fi),
                "Fecha final":   fmt_date(ff) if ff else "SIN FECHA",
                "Hora final":    fmt_time(ff) if ff else "",
                "Unidad":        str(b).strip(),
                "Potencia (MW)": float(g or 0),
                "Libranza":      str(e or "").strip(),
                "Descripción":   str(d or "").strip(),
                "status":        "vieja",
            })
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=INDISP_COLS+["status"])
    except Exception as e:
        return pd.DataFrame(columns=INDISP_COLS+["status"])

# ══════════════════════════════════════════════════════════════════════
# WEEK LOOKUP
# ══════════════════════════════════════════════════════════════════════
def find_week_for_date(target, weeks):
    """Find week number using Sábado/Viernes only (avoids bad Domingo year)."""
    if target is None: return None
    dt = parse_dt(target)
    if dt is None: return None
    target_d = dt.date() if isinstance(dt, datetime) else dt
    for wn, dates in weeks.items():
        sab = dates[0]   # Sábado — always correct year
        vie = dates[6]   # Viernes — always correct year
        if sab is None or vie is None: continue
        sab_d = sab.date() if isinstance(sab, datetime) else sab
        vie_d = vie.date() if isinstance(vie, datetime) else vie
        if sab_d <= target_d <= vie_d:
            return wn
    return None

# ══════════════════════════════════════════════════════════════════════
# PROCESSORS
# ══════════════════════════════════════════════════════════════════════
def process_nuevas(df_raw, lineas_codes=None, unit_mw=None, plant_prefix=None, equipos_db=None):
    """Filter and classify libranzas nuevas."""
    if df_raw.empty: return pd.DataFrame(columns=PLANTILLA_LIB_COLS)
    # Exclude ALL Cancelado + ALL Emergencia
    mask = (df_raw["Último Estado"].str.strip() != "Cancelado") & \
           (df_raw["Tipo"].str.strip() != "Emergencia")
    df = df_raw[mask].copy()

    out_rows = []
    for _, row in df.iterrows():
        ri = classify_libranza(row, lineas_codes, unit_mw, plant_prefix, equipos_db)
        out_rows.append({
            "Número":           row["Número"],
            "Tipo":             row["Tipo"],
            "Es repetitiva":    row["Es repetitiva"],
            "Fecha Solicitud":  row["Fecha Solicitud"],
            "Fecha Inicio":     row["Fecha Inicio"],
            "Fecha Final":      row["Fecha Final"],
            "Duración Programada": row["Duración Programada"],
            "Fecha Inicio Real":   row["Fecha Inicio Real"],
            "Fecha Final Real":    row["Fecha Final Real"],
            "Duración Real":       row["Duración Real"],
            "Descripción":         row["Descripción"],
            "Observaciones":       row["Observaciones"],
            "Equipos":             row["Equipos"],
            "Libranzas Vinculadas": row.get("Libranzas Vinculadas",""),
            "Fecha Aprobación":    row["Fecha Aprobación"],
            "Último Estado":       row["Último Estado"],
            "R/I":                 ri,
        })
    return pd.DataFrame(out_rows)

def process_viejas(df_raw, lineas_codes=None, unit_mw=None, plant_prefix=None, equipos_db=None):
    """Filter libranzas viejas: keep only Aprobado and Recibido, classify R/I."""
    if df_raw.empty: return pd.DataFrame(columns=PLANTILLA_LIB_COLS)
    keep = ["Aprobado","Recibido"]
    df = df_raw[df_raw["Último Estado"].str.strip().isin(keep)].copy()
    out_rows = []
    for _, row in df.iterrows():
        ri = classify_libranza(row, lineas_codes, unit_mw, plant_prefix, equipos_db)
        out_rows.append({
            "Número":           row["Número"],
            "Tipo":             row["Tipo"],
            "Es repetitiva":    row["Es repetitiva"],
            "Fecha Solicitud":  row["Fecha Solicitud"],
            "Fecha Inicio":     row["Fecha Inicio"],
            "Fecha Final":      row["Fecha Final"],
            "Duración Programada": row["Duración Programada"],
            "Fecha Inicio Real":   row["Fecha Inicio Real"],
            "Fecha Final Real":    row["Fecha Final Real"],
            "Duración Real":       row["Duración Real"],
            "Descripción":         row["Descripción"],
            "Observaciones":       row["Observaciones"],
            "Equipos":             row["Equipos"],
            "Libranzas Vinculadas": row.get("Libranzas Vinculadas",""),
            "Fecha Aprobación":    row["Fecha Aprobación"],
            "Último Estado":       row["Último Estado"],
            "R/I":                 ri,
        })
    return pd.DataFrame(out_rows)

def build_relevantes(df_nuevas, df_viejas, lineas_lookup, relevantes_anteriores=None, week_start=None):
    """Build relevantes: viejas first (blue), then nuevas (no color), sorted by Fecha inicio.
    Excludes: libranza informativa, secondary linked libranzas sharing de-energization."""
    rows      = []
    full_data = {}   # libranza_num → full source row
    seen      = set()
    viejas_nums = set(str(r) for r in df_viejas["Número"].dropna()) \
                  if df_viejas is not None and not df_viejas.empty else set()
    DEENERG_KW = ["desenergiz","de-energiz","desconex","apertura","seccionamiento"]

    def add_relevantes_from(df, classify=False, source_status="nueva"):
        if df is None or df.empty: return
        for _, row in df.iterrows():
            desc = str(row.get("Descripción","") or "").lower()
            if "libranza informativa" in desc or "lib. informativa" in desc:
                continue
            ri = row.get("R/I","") if not classify else classify_libranza(row)
            if "R" not in str(ri): continue
            eq_up = str(row.get("Equipos","") or "").upper()
            if "GENERADOR AUXILIAR" in eq_up and not any(
                t in eq_up for t in ["LINEA","AUTOTRANSFORMADOR","BANCO DE CAPACITORES"]):
                continue
            num = str(row.get("Número",""))
            if num in seen: continue
            seen.add(num)
            full_data[num] = row
            fi = parse_dt(row.get("Fecha Inicio",""))
            ff = parse_dt(row.get("Fecha Final",""))
            info = get_relevante_equipo_info(row, lineas_lookup)
            rows.append({
                "Tipo":               row.get("Es repetitiva",""),
                "Fecha inicio":       fmt_date(fi),
                "Hora inicio":        fmt_time(fi),
                "Fecha final":        fmt_date(ff),
                "Hora final":         fmt_time(ff),
                "Tipo de Equipos":    info["tipo"],
                "Equipo":             info["equipo"],
                "Subestación":        info["sust"],
                "Libranza":           num,
                "Descripción del trabajo": row.get("Descripción",""),
                "Estado":             row.get("Último Estado",""),
                "Observaciones":      row.get("Observaciones",""),
                "_sort_dt":           fi or datetime.max,
                "_status":            source_status,
            })

    add_relevantes_from(df_viejas, classify=False, source_status="vieja")
    add_relevantes_from(df_nuevas, classify=False, source_status="nueva")

    if relevantes_anteriores:
        for r in relevantes_anteriores:
            lib = r.get("Libranza","")
            if not lib or lib in seen: continue
            ff = parse_dt(r.get("Fecha final",""))
            if week_start and ff and ff.date() < week_start: continue
            seen.add(lib)
            status = "vieja" if lib in viejas_nums else "nueva"
            rows.append({k: r[k] for k in PLANTILLA_REL_COLS} | {
                "_sort_dt": r.get("_sort_dt", datetime.max),
                "_status":  status,
            })

    # Filter secondary linked libranzas (vinculada is base → remove secondary)
    rel_nums  = {r["Libranza"] for r in rows}
    to_remove = set()
    for r in rows:
        num     = r["Libranza"]
        src     = full_data.get(num)
        if src is None: continue
        vinc    = str(src.get("Libranzas Vinculadas","") or "")
        if not vinc or vinc in ("None","nan","No hay libranzas vinculadas",""): continue
        for vin in re.split(r"[,;\n\s]+", vinc):
            vin = vin.strip()
            if not vin or vin not in rel_nums: continue
            vin_src = full_data.get(vin)
            if vin_src is None: continue
            fi_curr = parse_dt(src.get("Fecha Inicio",""))
            ff_curr = parse_dt(src.get("Fecha Final",""))
            fi_vin  = parse_dt(vin_src.get("Fecha Inicio",""))
            ff_vin  = parse_dt(vin_src.get("Fecha Final",""))
            if fi_curr is None or fi_vin is None: continue
            # Vinculada (base) started before current AND current dates within base range
            if fi_vin <= fi_curr:
                dates_within = (ff_vin is None or ff_curr is None or
                                fi_curr >= fi_vin and (ff_curr <= ff_vin or abs((ff_curr - ff_vin).total_seconds()) < 3600))
                eq_c = extract_equip_codes(str(src.get("Equipos","") or ""))
                eq_v = extract_equip_codes(str(vin_src.get("Equipos","") or ""))
                if dates_within and eq_c & eq_v:
                    to_remove.add(num)
    rows = [r for r in rows if r["Libranza"] not in to_remove]

    if not rows: return pd.DataFrame(columns=PLANTILLA_REL_COLS + ["_status"])
    df_out = pd.DataFrame(rows)
    df_out["_is_nueva"] = (df_out["_status"] == "nueva").astype(int)
    df_out = df_out.sort_values(["_is_nueva","_sort_dt"]).drop(columns=["_is_nueva","_sort_dt"])
    return df_out.reset_index(drop=True)

def build_indisponibilidades(indisp_existing, df_viejas, df_nuevas,
                             unit_mw, plant_prefix, weeks, current_week,
                             indisp_file_df=None):
    """Build flat indisponibilidades DataFrame.
    Priority: external file > plantilla existing > nuevas-I
    status: 'vieja' (blue) | 'nueva' (white)
    """
    week_dates = weeks.get(current_week, [None]*7)
    sab = week_dates[0]
    fixed_dates = []
    for i, d in enumerate(week_dates):
        if d is None: fixed_dates.append(None)
        elif sab and abs((d - sab).days) > 7:
            fixed_dates.append(sab + timedelta(days=i))
        else: fixed_dates.append(d)
    valid_dts = [d for d in fixed_dates if d]
    if not valid_dts: return pd.DataFrame(columns=INDISP_COLS + ["status"])

    week_start = valid_dts[0].date()
    week_end   = valid_dts[-1].date()
    result = []
    seen   = set()

    # ── 1. External indisponibilidades file (primary source) ───────
    if indisp_file_df is not None and not indisp_file_df.empty:
        for _, r in indisp_file_df.iterrows():
            fi = parse_dt(r.get("Fecha inicio",""))
            ff_str = str(r.get("Fecha final",""))
            ff = parse_dt(ff_str) if ff_str not in ("SIN FECHA","") else None
            lib = str(r.get("Libranza","")).strip()
            if fi is None: continue
            fi_d = fi.date()
            ff_d = ff.date() if ff else date(2099, 12, 31)
            if fi_d <= week_end and ff_d >= week_start:
                entry = {k: r.get(k,"") for k in INDISP_COLS}
                entry["status"] = "vieja"
                result.append(entry)
                if lib: seen.add(lib)
    else:
        # ── 2. Fallback: plantilla existing filtered by viejas ──────
        viejas_nums = set(str(r) for r in df_viejas["Número"].dropna()) \
                      if df_viejas is not None and not df_viejas.empty else set()
        for entry in (indisp_existing or []):
            lib = str(entry.get("Libranza",""))
            if not lib or lib in seen: continue
            if lib in viejas_nums:
                result.append({**{k: entry.get(k,"") for k in INDISP_COLS}, "status": "vieja"})
            seen.add(lib)

    # ── 3. New I entries from libranzas_nuevas ─────────────────────
    GEN_TYPES = ["GENERADOR","UNIDAD DE GENERACION","GRUPO GENERADOR","INVERSOR ESTATICO"]
    if df_nuevas is not None and not df_nuevas.empty:
        df_i = df_nuevas[df_nuevas["R/I"].str.contains("I", na=False)]
        for _, row in df_i.iterrows():
            eq = str(row.get("Equipos","") or "").upper()
            if not any(t in eq for t in GEN_TYPES): continue
            num = str(row.get("Número",""))
            if num in seen: continue
            seen.add(num)
            fi = parse_dt(row.get("Fecha Inicio",""))
            ff = parse_dt(row.get("Fecha Final",""))
            if fi is None or ff is None: continue
            unit_label, mw = get_unit_name_and_mw(row, unit_mw or {}, plant_prefix or {})
            result.append({
                "Fecha inicio": fmt_date(fi), "Hora inicio": fmt_time(fi),
                "Fecha final":  fmt_date(ff), "Hora final":  fmt_time(ff),
                "Unidad": unit_label, "Potencia (MW)": mw,
                "Libranza": num, "Descripción": str(row.get("Descripción","")),
                "status": "nueva"
            })

    if not result: return pd.DataFrame(columns=INDISP_COLS + ["status"])
    df = pd.DataFrame(result)
    df["_ff"] = df["Fecha final"].apply(parse_dt)
    df_v = df[df["status"]=="vieja"].copy()
    df_n = df[df["status"]=="nueva"].sort_values("_ff")
    return pd.concat([df_v, df_n], ignore_index=True).drop(columns=["_ff"])


def update_proyectos(df_proy, current_week, weeks):
    """Update Sem Disp (from fecha_fin lookup) and Sem Prueba (replace week number)."""
    if df_proy.empty: return df_proy
    df = df_proy.copy()
    prev_week = current_week - 1
    for idx, row in df.iterrows():
        raw = row.get("_fecha_fin_raw")
        fecha_str = row.get("Fecha finaliza libranza","")
        dt = parse_dt(raw) if (raw is not None and str(raw) not in ["None","nan",""]) \
             else parse_dt(fecha_str)
        if dt is not None:
            wn = find_week_for_date(dt, weeks)
            if wn is not None:
                df.at[idx,"Sem Disp"] = f"Semana {wn}"
        sem_p = str(row.get("Sem Prueba","") or "")
        if sem_p and str(prev_week) in sem_p:
            df.at[idx,"Sem Prueba"] = re.sub(
                rf'\bSemana\s+{prev_week}\b', f'Semana {current_week}', sem_p)
    return df

def detect_proyectos_from_libranzas(df_viejas, df_nuevas, df_proy_existing):
    """Detect libranzas with 'prueba de generacion' not yet in proyectos, add them."""
    PRUEBA_KW = ["prueba de generacion", "prueba de generación", "pruebas de generacion",
                 "pruebas de generación", "prueba de puesta en servicio"]
    existing_libs = set(str(r).strip() for r in df_proy_existing["Última Solicitud"].dropna()) \
                    if not df_proy_existing.empty else set()

    new_rows = []
    for df in [df_viejas, df_nuevas]:
        if df is None or df.empty: continue
        for _, row in df.iterrows():
            num = str(row.get("Número",""))
            desc = str(row.get("Descripción","") or "").lower()
            if not any(kw in desc for kw in PRUEBA_KW): continue
            if num in existing_libs: continue  # already in proyectos
            # Try to extract plant name from Equipos
            equipos = str(row.get("Equipos","") or "")
            m = re.search(r'(?:PLANTA|SE)\s+([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñ\s]+?)\s*->', equipos)
            planta = m.group(1).strip().title() if m else num
            ff = row.get("Fecha Final","")
            new_rows.append({
                "Planta":     planta,
                "Tecnología": "",
                "MW":         None,
                "Fecha Planeamiento": "",
                "Sem Disp":   "",
                "Sem Prueba": "",
                "Última Solicitud": num,
                "Fecha finaliza libranza": ff,
                "_fecha_fin_raw": parse_dt(str(ff)),
            })
            existing_libs.add(num)

    if not new_rows: return df_proy_existing
    df_new = pd.DataFrame(new_rows)
    return pd.concat([df_proy_existing, df_new], ignore_index=True)
    """Update Sem Disp (from fecha_fin lookup) and Sem Prueba (replace week number)."""
    if df_proy.empty: return df_proy
    df = df_proy.copy()
    prev_week = current_week - 1

    for idx, row in df.iterrows():
        # ── Sem Disp: look up fecha_fin in calendar ──────────────────
        raw = row.get("_fecha_fin_raw")
        fecha_str = row.get("Fecha finaliza libranza","")
        dt = parse_dt(raw) if (raw is not None and str(raw) not in ["None","nan",""]) \
             else parse_dt(fecha_str)
        if dt is not None:
            wn = find_week_for_date(dt, weeks)
            if wn is not None:
                df.at[idx,"Sem Disp"] = f"Semana {wn}"

        # ── Sem Prueba: replace previous week number with current ─────
        sem_p = str(row.get("Sem Prueba","") or "")
        if sem_p and str(prev_week) in sem_p:
            df.at[idx,"Sem Prueba"] = re.sub(
                rf'\bSemana\s+{prev_week}\b',
                f'Semana {current_week}',
                sem_p
            )
    return df

# ══════════════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════════════
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
TITLE_FONT = Font(bold=True, color="1F3864", size=11)
DATA_FONT  = Font(size=9)
THIN  = Side(border_style="thin", color="CCCCCC")
BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP  = Alignment(vertical="top", wrap_text=True)
CTR   = Alignment(horizontal="center", vertical="center")

def _hdr(ws, headers, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CTR; cell.border = BORD

def _write(ws, df, start_row=3):
    for ri, (_, r) in enumerate(df.iterrows(), start_row):
        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(ri, ci, r[col])
            cell.font = DATA_FONT; cell.border = BORD; cell.alignment = WRAP

def _clear_rows(ws, min_row):
    """Clear data rows safely — values, fills and borders."""
    _wf = PatternFill("solid", fgColor="FFFFFF")
    _nb = Border()
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row):
        for cell in row:
            try:
                cell.value = None; cell.fill = _wf; cell.border = _nb
            except AttributeError: pass

def export_premisas(state):
    """Generate updated plantilla Excel (without Lineas, Datos, LIBRANZAS NUEVAS/VIEJAS)."""
    wb_template = load_workbook(BytesIO(state["prem_plantilla_bytes"]))

    # ── Remove unwanted sheets ────────────────────────────────────────
    for sheet_name in ["Lineas", "Datos", "LIBRANZAS NUEVAS", "LIBRANZAS VIEJAS"]:
        if sheet_name in wb_template.sheetnames:
            del wb_template[sheet_name]

    # ── LIBRANZAS RELEVANTES ─────────────────────────────────────────
    ws = wb_template["Libranzas Relevantes"]
    for merge in list(ws.merged_cells.ranges): ws.unmerge_cells(str(merge))
    _clear_rows(ws, 3)
    _hdr(ws, PLANTILLA_REL_COLS, row=2)
    df_r = state.get("prem_df_relevantes", pd.DataFrame())
    if not df_r.empty:
        BLUE_FILL_R  = PatternFill("solid", fgColor="BDD7EE")   # blue for viejas
        GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")   # green for Continua
        PINK_FILL    = PatternFill("solid", fgColor="FFB6C1")   # pink for Repetitiva
        WHITE_FILL_R = PatternFill("solid", fgColor="FFFFFF")

        for ri, (_, r) in enumerate(df_r.iterrows(), 3):
            status = str(r.get("_status","")) if "_status" in r.index else ""
            is_vieja = status == "vieja"
            tipo_val = str(r.get("Tipo","")).strip().lower()

            for ci, col in enumerate(PLANTILLA_REL_COLS, 1):
                val = r.get(col,"") if col in r.index else ""
                try:
                    cell = ws.cell(ri, ci, val)
                    cell.font      = DATA_FONT
                    cell.border    = BORD
                    cell.alignment = WRAP
                    cell.fill      = WHITE_FILL_R

                    # Tipo col (ci=1): green=Continua, pink=Repetitiva
                    if ci == 1:
                        if "continua" in tipo_val:
                            cell.fill = GREEN_FILL
                        elif "repetitiva" in tipo_val:
                            cell.fill = PINK_FILL

                    # Blue ONLY on Libranza column (ci=9) for viejas
                    elif ci == 9 and is_vieja:
                        cell.fill = BLUE_FILL_R
                        cell.font = Font(size=9, bold=True)
                except: pass

    # ── INDISPONIBILIDADES ───────────────────────────────────────────
    BLUE_FILL  = PatternFill("solid", fgColor="BDD7EE")   # blue for viejas
    TOTAL_FONT = Font(bold=True, size=9)

    ws = wb_template["Indisponibilidades"]
    cur_week = state.get("prem_current_week", 1)
    for merge in list(ws.merged_cells.ranges): ws.unmerge_cells(str(merge))
    try: ws.cell(1, 1).value = f"Indisponibilidades de Generación Intersemanales - Semana {cur_week}"
    except: pass
    try: ws.cell(2, 1).value = f"Semana {cur_week}"
    except: pass
    _clear_rows(ws, 3)

    indisp_df  = state.get("prem_indisp_data")
    weeks      = state.get("prem_weeks", {})
    week_dates = weeks.get(cur_week, [None]*7)

    # Fix Domingo dates
    sab = week_dates[0] if week_dates else None
    fixed_dates = []
    for i, d in enumerate(week_dates):
        if d is None: fixed_dates.append(None)
        elif sab and abs((d-sab).days) > 7: fixed_dates.append(sab + timedelta(days=i))
        else: fixed_dates.append(d)

    ri = 3
    weekly_total = 0.0

    for i, dia in enumerate(DIAS_SEMANA):
        dt = fixed_dates[i] if i < len(fixed_dates) else None
        if dt is None: continue
        day_date = dt.date()

        # Filter indisp_df for this day
        if indisp_df is None or (hasattr(indisp_df,'empty') and indisp_df.empty):
            day_rows = []
        else:
            day_rows = []
            for _, row_data in indisp_df.iterrows():
                fi = parse_dt(row_data.get("Fecha inicio",""))
                ff = parse_dt(row_data.get("Fecha final",""))
                if fi is None or ff is None: continue
                if fi.date() <= day_date <= ff.date():
                    day_rows.append(row_data)

        if not day_rows: continue

        # Sort: viejas first, then nuevas; within each group by Fecha inicio ascending
        def sort_key(r):
            is_nueva = 1 if str(r.get("status","")) == "nueva" else 0
            fi = parse_dt(r.get("Fecha inicio","")) or datetime.max
            return (is_nueva, fi)
        day_rows = sorted(day_rows, key=sort_key)

        WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
        NO_BORDER  = Border()
        DATA_BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        day_start_ri = ri   # first data row for this day
        day_total    = 0.0

        for row_data in day_rows:
            status   = str(row_data.get("status","")) if hasattr(row_data, 'get') else ""
            is_vieja = status == "vieja"

            for ci, col in enumerate(INDISP_COLS, 2):
                val = row_data.get(col, "") if hasattr(row_data,'get') else ""
                try:
                    cell = ws.cell(ri, ci, val)
                    cell.font   = DATA_FONT
                    cell.border = DATA_BORD
                    cell.fill   = WHITE_FILL
                    # Blue ONLY on Libranza col (ci=8) for viejas
                    if is_vieja and ci == 8:
                        cell.fill = BLUE_FILL
                        cell.font = Font(size=9, bold=True)
                except: pass

            mw = float(row_data.get("Potencia (MW)", 0) or 0) if hasattr(row_data,'get') else 0
            day_total   += mw
            weekly_total += mw
            ri += 1

        # Merge col A for all data rows of this day
        day_end_ri = ri - 1
        try:
            if day_end_ri > day_start_ri:
                ws.merge_cells(f"A{day_start_ri}:A{day_end_ri}")
            cell_a = ws.cell(day_start_ri, 1)
            cell_a.value     = f"{dia.lower()} {day_date.day}"
            cell_a.font      = Font(size=9, bold=True)
            cell_a.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell_a.border    = DATA_BORD
            cell_a.fill      = WHITE_FILL
        except: pass

        # Daily total: merge B:F, value + MW in G, NO borders
        try:
            ws.merge_cells(f"B{ri}:F{ri}")
            cl = ws.cell(ri, 2)
            cl.value     = "INDISPONIBILIDAD TOTAL EN HORAS PUNTA (MW):"
            cl.font      = TOTAL_FONT
            cl.fill      = WHITE_FILL
            cl.border    = NO_BORDER
            cl.alignment = Alignment(vertical="center")
            cm = ws.cell(ri, 7)
            cm.value  = round(day_total, 4)
            cm.font   = TOTAL_FONT
            cm.fill   = WHITE_FILL
            cm.border = NO_BORDER
            for ci in [1, 3, 4, 5, 6, 8, 9]:
                try:
                    c = ws.cell(ri, ci)
                    c.value = None; c.fill = WHITE_FILL; c.border = NO_BORDER
                except: pass
        except: pass
        ri += 2

    # Weekly grand total
    try:
        WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
        ws.merge_cells(f"B{ri}:F{ri}")
        cl = ws.cell(ri, 2)
        cl.value = "INDISPONIBILIDAD TOTAL EN HORAS PUNTA (MW):"
        cl.font  = Font(bold=True, size=9, color="FF0000")
        cl.fill  = WHITE_FILL; cl.border = Border()
        cm = ws.cell(ri, 7)
        cm.value = round(weekly_total, 4)
        cm.font  = Font(bold=True, size=9, color="FF0000")
        cm.fill  = WHITE_FILL; cm.border = Border()
    except: pass

    # Clear everything below the grand total (remove old borders/fills)
    last_written = ri
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    for row in ws.iter_rows(min_row=last_written+1, max_row=min(last_written+200, ws.max_row)):
        for cell in row:
            try:
                cell.value  = None
                cell.fill   = WHITE_FILL
                cell.border = Border()
            except: pass

    # ── PROYECTOS DE GENERACIÓN ──────────────────────────────────────
    ws = wb_template["Proyectos de Generacion"]
    df_p = state.get("prem_df_proyectos", pd.DataFrame())
    if not df_p.empty:
        # Find data start row
        data_start = None
        for ridx, row in enumerate(ws.iter_rows(min_row=13, values_only=True), 13):
            if row[0] == "Planta":
                data_start = ridx + 1
                break
        if data_start:
            # Clear existing data
            _clear_rows(ws, data_start)
            for ri2, (_, r) in enumerate(df_p.iterrows(), data_start):
                cols_to_write = [c for c in PROY_COLS if c in df_p.columns]
                for ci, col in enumerate(cols_to_write, 1):
                    ws.cell(ri2, ci).value = r.get(col,"")
                    ws.cell(ri2, ci).font = DATA_FONT

    buf = BytesIO()
    wb_template.save(buf)
    buf.seek(0)
    return buf.read()

# ══════════════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════════════
def vista_premisas():
    st.markdown("## ⚡ Módulo de Premisas")

    # ── Session state init ────────────────────────────────────────────
    for key in ["plantilla","prem_df_nuevas","prem_df_viejas","prem_df_relevantes",
                "prem_indisp_data","prem_df_proyectos","prem_current_week","prem_weeks",
                "prem_lineas_lookup","prem_plantilla_bytes"]:
        if key not in st.session_state:
            st.session_state[key] = None

    # ══════════════════════════════════════════════════════════════════
    # SIDEBAR
    # ══════════════════════════════════════════════════════════════════
    with st.sidebar:
        st.title("📂 Archivos")
        st.divider()

        f_plantilla = st.file_uploader("1. Plantilla semana anterior", type="xlsx", key="prem_up_plantilla")
        f_nuevas    = st.file_uploader("2. libranzas_nuevas.xlsx",     type="xlsx", key="prem_up_nuevas")
        f_viejas    = st.file_uploader("3. libranzas_viejas.xlsx",     type="xlsx", key="prem_up_viejas")
        f_indisp    = st.file_uploader("4. Indisponibilidades (opcional)", type="xlsx", key="prem_up_indisp")
        # equipos_libranzas.xlsx loaded automatically from repo

        st.divider()
        btn_process = st.button("⚙️ Procesar", type="primary", use_container_width=True,
                                disabled=not (f_plantilla and f_nuevas and f_viejas))

        if btn_process:
            with st.spinner("Procesando..."):
                try:
                    plantilla_bytes = f_plantilla.read()
                    pl = load_plantilla(plantilla_bytes)
                    df_src_nuevas = load_source_libranzas(f_nuevas.read())
                    df_src_viejas = load_source_libranzas(f_viejas.read())
                    indisp_file_df = load_indisp_file(f_indisp.read()) if f_indisp else None
                    import os as _os
                    _eq_path = "equipos_libranzas.xlsx"
                    equipos_db = load_equipos_db(open(_eq_path,"rb").read()) if _os.path.exists(_eq_path) else None

                    current_week  = pl["current_week"]
                    weeks         = pl["weeks"]
                    lineas        = pl["lineas_lookup"]
                    rel_ant       = pl.get("relevantes_anteriores", [])
                    unit_mw       = pl.get("unit_mw", {})
                    plant_prefix  = pl.get("plant_prefix", {})
                    lineas_codes  = pl.get("lineas_codes", set())
                    indisp_exist  = pl.get("indisp_existing", [])

                    # Week start date
                    week_dates    = weeks.get(current_week, [None]*7)
                    sab = week_dates[0]
                    fixed = [sab + timedelta(days=i) if (d and sab and abs((d-sab).days)>7) else d
                             for i,d in enumerate(week_dates)]
                    week_start = next((d.date() for d in fixed if d), None)

                    df_nuevas    = process_nuevas(df_src_nuevas, lineas_codes, unit_mw, plant_prefix, equipos_db)
                    df_viejas    = process_viejas(df_src_viejas, lineas_codes, unit_mw, plant_prefix, equipos_db)
                    df_relevantes= build_relevantes(df_nuevas, df_viejas, lineas,
                                                    rel_ant, week_start)
                    indisp_data  = build_indisponibilidades(
                                        indisp_exist, df_viejas, df_nuevas,
                                        unit_mw, plant_prefix, weeks, current_week,
                                        indisp_file_df)
                    df_proyectos = update_proyectos(pl["proyectos"], current_week, weeks)
                    df_proyectos = detect_proyectos_from_libranzas(df_viejas, df_nuevas, df_proyectos)

                    st.session_state.update({
                        "prem_plantilla_bytes": plantilla_bytes,
                        "prem_current_week":    current_week,
                        "prem_weeks":           weeks,
                        "prem_lineas_lookup":   lineas,
                        "prem_unit_mw":         unit_mw,
                        "prem_plant_prefix":    plant_prefix,
                        "prem_lineas_codes":    lineas_codes,
                        "prem_equipos_db":      equipos_db,
                        "prem_df_nuevas":       df_nuevas,
                        "prem_df_viejas":       df_viejas,
                        "prem_df_relevantes":   df_relevantes,
                        "prem_indisp_data":     indisp_data,
                        "prem_df_proyectos":    df_proyectos,
                        "prem_procesado":       True,
                        "prem_relevantes_anteriores": rel_ant,
                        "prem_week_start":      week_start,
                    })
                    st.success(f"✅ Semana {current_week} procesada")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")
                    import traceback; st.text(traceback.format_exc())

        if st.session_state.prem_current_week:
            st.divider()
            cw = st.session_state.prem_current_week
            wdates = st.session_state.prem_weeks.get(cw,[]) if st.session_state.prem_weeks else []
            st.markdown(f'<div class="week-badge">Semana {cw}</div>', unsafe_allow_html=True)
            st.caption(week_range(wdates))
            c1,c2 = st.columns(2)
            df_n = st.session_state.prem_df_nuevas
            df_v = st.session_state.prem_df_viejas
            c1.metric("Nuevas",  len(df_n) if df_n is not None else 0)
            c2.metric("Viejas",  len(df_v) if df_v is not None else 0)

            st.divider()
            if st.session_state.prem_df_nuevas is not None:
                if st.button("📦 Preparar exportación", use_container_width=True):
                    with st.spinner("Generando archivo..."):
                        st.session_state.prem_export_bytes = export_premisas({
                            "prem_plantilla_bytes": st.session_state.prem_plantilla_bytes,
                            "prem_current_week":    st.session_state.prem_current_week,
                            "prem_weeks":           st.session_state.prem_weeks,
                            "prem_df_nuevas":       st.session_state.prem_df_nuevas,
                            "prem_df_viejas":       st.session_state.prem_df_viejas,
                            "prem_df_relevantes":   st.session_state.prem_df_relevantes,
                            "prem_indisp_data":   st.session_state.prem_indisp_data,
                            "prem_df_proyectos":    st.session_state.prem_df_proyectos,
                        })
                if st.session_state.get("prem_export_bytes"):
                    st.download_button(
                        "📥 Descargar plantilla",
                        data=st.session_state.prem_export_bytes,
                        file_name=f"Premisas_SEM_{cw:02d}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

    # ══════════════════════════════════════════════════════════════════
    # MAIN CONTENT
    # ══════════════════════════════════════════════════════════════════
    if st.session_state.prem_df_nuevas is None:
        st.info("👈 Cargue los archivos en el panel lateral y presione **Procesar** para comenzar.")
        return

    cw      = st.session_state.prem_current_week
    weeks   = st.session_state.prem_weeks
    wdates  = weeks.get(cw, [None]*7)

    tabs = st.tabs([
        "🆕 Libranzas Nuevas",
        "📋 Libranzas Viejas",
        "⚠️ Indisponibilidades",
        "🔑 Relevantes",
        "🏭 Proyectos",
    ])

    # ── TAB 0: Libranzas Nuevas ───────────────────────────────────────
    with tabs[0]:
        st.subheader(f"Libranzas Nuevas — Semana {cw}")
        df = st.session_state.prem_df_nuevas.copy()

        total_r  = len(df[df["R/I"].str.contains("R", na=False)])
        total_i  = len(df[df["R/I"].str.contains("I", na=False)])
        st.info(f"✅ Filtrado automático aplicado — **{len(df)} libranzas** | "
                f"**{total_r} Relevantes (R)** | **{total_i} Indisponibilidades (I)**")

        buscar = st.text_input("🔍 Buscar", key="prem_search_nuevas", placeholder="Ej: ETESA-615")
        if buscar:
            df = df[df["Número"].astype(str).str.contains(buscar, case=False, na=False)]

        disp_cols_want = ["Número","Es repetitiva","Fecha Inicio","Fecha Final","Equipos","R/I","Descripción"]
        disp_cols = [c for c in disp_cols_want if c in df.columns]
        st.caption(f"{len(df)} libranzas")

        edited = st.data_editor(
            df[disp_cols].reset_index(drop=True),
            use_container_width=True, hide_index=True,
            num_rows="dynamic", key="prem_editor_nuevas",
            height=500,
            column_config={
                "R/I": st.column_config.SelectboxColumn("R/I", options=["","R","I","R-I"], width="small"),
                "Número": st.column_config.TextColumn("Número", width="medium"),
                "Es repetitiva": st.column_config.TextColumn("Tipo", width="small"),
                "Fecha Inicio": st.column_config.TextColumn("Fecha Inicio", width="medium"),
                "Fecha Final": st.column_config.TextColumn("Fecha Final", width="medium"),
                "Equipos": st.column_config.TextColumn("Equipos", width="large"),
                "Descripción": st.column_config.TextColumn("Descripción", width="large"),
            }
        )
        if st.button("💾 Guardar cambios", key="prem_apply_nuevas"):
            full = st.session_state.prem_df_nuevas.copy()
            if buscar:
                mask = full["Número"].astype(str).str.contains(buscar, case=False, na=False)
                vals = edited["R/I"].values if "R/I" in edited.columns else []
                if len(vals) == mask.sum():
                    full.loc[mask, "R/I"] = vals
            else:
                if "R/I" in edited.columns:
                    full["R/I"] = edited["R/I"].values
            st.session_state.prem_df_nuevas = full
            st.success(f"✅ Guardado — {len(full)} libranzas")

    # ── TAB 1: Libranzas Viejas ───────────────────────────────────────
    with tabs[1]:
        st.subheader(f"Libranzas Viejas — Semana {cw}")
        df_v = st.session_state.prem_df_viejas.copy()

        n_r_v = df_v["R/I"].str.contains("R", na=False).sum() if "R/I" in df_v.columns else 0
        n_i_v = df_v["R/I"].str.contains("I", na=False).sum() if "R/I" in df_v.columns else 0
        st.info(f"✅ Filtrado automático aplicado — **{len(df_v)} libranzas** con estado Aprobado o Recibido | "
                f"**{n_r_v} Relevantes (R)** | **{n_i_v} Indisponibilidades (I)**")

        buscar_v = st.text_input("🔍 Buscar", key="prem_search_viejas", placeholder="Ej: ETESA-571")
        if buscar_v:
            df_v = df_v[df_v["Número"].astype(str).str.contains(buscar_v, case=False, na=False)]

        disp_v_want = ["Número","Es repetitiva","Fecha Inicio","Fecha Final","Equipos","R/I","Descripción"]
        disp_v = [c for c in disp_v_want if c in df_v.columns]
        if not disp_v:
            st.warning(f"Columnas disponibles: {list(df_v.columns)}")
            disp_v = list(df_v.columns)
        st.caption(f"{len(df_v)} libranzas")
        edited_v = st.data_editor(
            df_v[disp_v].reset_index(drop=True),
            use_container_width=True, hide_index=True,
            num_rows="dynamic", key="prem_editor_viejas", height=500,
            column_config={
                "R/I": st.column_config.SelectboxColumn("R/I", options=["","R","I","R-I"], width="small"),
                "Es repetitiva": st.column_config.TextColumn("Tipo", width="small"),
                "Equipos": st.column_config.TextColumn("Equipos", width="large"),
                "Descripción": st.column_config.TextColumn("Descripción", width="large"),
            }
        )
        if st.button("💾 Guardar cambios", key="prem_apply_viejas"):
            full_v = st.session_state.prem_df_viejas.copy()
            if buscar_v:
                mask_v = full_v["Número"].astype(str).str.contains(buscar_v, case=False, na=False)
                if "R/I" in edited_v.columns and len(edited_v) == mask_v.sum():
                    full_v.loc[mask_v, "R/I"] = edited_v["R/I"].values
            else:
                if "R/I" in edited_v.columns:
                    full_v["R/I"] = edited_v["R/I"].values
            st.session_state.prem_df_viejas = full_v
            st.success(f"✅ Guardado — {len(full_v)} libranzas")

    # ── TAB 2: Indisponibilidades ─────────────────────────────────────
    with tabs[2]:
        st.subheader(f"Indisponibilidades — Semana {cw}")
        indisp_data = st.session_state.get("prem_indisp_data")
        if indisp_data is None or (hasattr(indisp_data,'empty') and indisp_data.empty):
            st.info("No hay indisponibilidades registradas.")
        else:
            df_id = indisp_data.copy()
            n_v = (df_id["status"]=="vieja").sum()
            n_n = (df_id["status"]=="nueva").sum()
            total_hp = pd.to_numeric(df_id["Potencia (MW)"], errors="coerce").sum()
            st.info(f"**{len(df_id)}** — {n_v} anteriores | {n_n} nuevas | Total HP: **{total_hp:.2f} MW**")
            buscar_id = st.text_input("🔍 Buscar", key="prem_search_indisp",
                                      placeholder="Ej: CELSIACENT o BAYG3")
            df_id_f = df_id.copy()
            if buscar_id:
                m = df_id_f.apply(lambda r: buscar_id.lower() in
                    " ".join(str(v) for v in r.values).lower(), axis=1)
                df_id_f = df_id_f[m]

            # Build week dates
            wd = weeks.get(cw, [None]*7); sab = wd[0]
            fd = [sab+timedelta(days=i) if (d and sab and abs((d-sab).days)>7) else d
                  for i,d in enumerate(wd)]
            grand = 0.0

            for i_d, dia in enumerate(DIAS_SEMANA):
                dt = fd[i_d] if i_d < len(fd) else None
                if dt is None: continue
                dd = dt.date()
                rows_d = df_id_f[df_id_f.apply(
                    lambda r: (lambda a,b: bool(a and b and a.date()<=dd<=b.date()))(
                        parse_dt(r.get("Fecha inicio","")), parse_dt(r.get("Fecha final",""))), axis=1)]
                if rows_d.empty: continue
                dtot = pd.to_numeric(rows_d["Potencia (MW)"], errors="coerce").sum()
                grand += dtot
                st.markdown(f"**{dia.lower()} {dd.day}** &nbsp;·&nbsp; Total HP: `{dtot:.2f} MW`",
                            unsafe_allow_html=True)

                # Styled preview (colored like Excel)
                dcols = [c for c in ["Unidad","Fecha inicio","Hora inicio","Fecha final",
                                     "Hora final","Potencia (MW)","Libranza","Descripción"]
                         if c in rows_d.columns]
                df_styled = rows_d[dcols].reset_index(drop=True)
                sts = rows_d["status"].tolist() if "status" in rows_d.columns else []
                def _sty(row, _s=sts):
                    s = _s[row.name] if row.name < len(_s) else ""
                    if s == "vieja":
                        return ["background-color:#BDD7EE" if c=="Libranza" else "" for c in row.index]
                    return [""]*len(row)
                try:
                    st.dataframe(df_styled.style.apply(_sty, axis=1),
                                 use_container_width=True, hide_index=True)
                except:
                    st.dataframe(df_styled, use_container_width=True, hide_index=True)

            st.markdown(f"**INDISPONIBILIDAD TOTAL EN HORAS PUNTA (MW): `{grand:.2f}`**")

            # Editable section
            with st.expander("✏️ Editar indisponibilidades"):
                disp_e = [c for c in ["Unidad","Fecha inicio","Hora inicio","Fecha final",
                                      "Hora final","Potencia (MW)","Libranza","Descripción","status"]
                          if c in df_id.columns]
                ed_id = st.data_editor(
                    df_id[disp_e].reset_index(drop=True),
                    use_container_width=True, hide_index=True, num_rows="dynamic",
                    key="prem_editor_indisp",
                    column_config={
                        "status": st.column_config.SelectboxColumn(
                            "Estado", options=["vieja","nueva"], width="small"),
                        "Potencia (MW)": st.column_config.NumberColumn(format="%.2f"),
                    }
                )
                if st.button("💾 Guardar", key="prem_apply_indisp"):
                    st.session_state.prem_indisp_data = ed_id.copy()
                    st.success("✅ Guardado"); st.rerun()

            # Editable section
            with st.expander("✏️ Editar indisponibilidades"):
                disp_e = [c for c in ["Unidad","Fecha inicio","Hora inicio","Fecha final",
                                      "Hora final","Potencia (MW)","Libranza","Descripción","status"]
                          if c in df_id.columns]
                ed_id = st.data_editor(
                    df_id[disp_e].reset_index(drop=True),
                    use_container_width=True, hide_index=True, num_rows="dynamic",
                    key="prem_editor_indisp",
                    column_config={
                        "status": st.column_config.SelectboxColumn(
                            "Estado", options=["vieja","nueva"], width="small"),
                        "Potencia (MW)": st.column_config.NumberColumn(format="%.2f"),
                    }
                )
                if st.button("💾 Guardar", key="prem_apply_indisp"):
                    st.session_state.prem_indisp_data = ed_id.copy()
                    st.success("✅ Guardado"); st.rerun()

    # ── TAB 3: Libranzas Relevantes ───────────────────────────────────
    with tabs[3]:
        st.subheader(f"Libranzas Relevantes — Semana {cw}")
        df_rel = (st.session_state.prem_df_relevantes.copy()
                  if st.session_state.prem_df_relevantes is not None
                  else pd.DataFrame(columns=PLANTILLA_REL_COLS))
        if df_rel.empty:
            st.info("No hay libranzas clasificadas como R.")
        else:
            st.info(f"**{len(df_rel)} relevantes**")
            buscar_r = st.text_input("🔍 Buscar", key="prem_search_rel",
                                     placeholder="Ej: ETESA-688 o 230-4A")
            disp_r = [c for c in ["Tipo","Fecha inicio","Fecha final","Tipo de Equipos",
                                   "Equipo","Subestación","Libranza","Descripción del trabajo","Estado"]
                      if c in df_rel.columns]
            df_rs = df_rel[disp_r].copy() if disp_r else df_rel.copy()
            sts_r = df_rel["_status"].tolist() if "_status" in df_rel.columns else [""] * len(df_rel)

            if buscar_r:
                mask_r = df_rs.apply(lambda row: buscar_r.lower() in
                    " ".join(str(v) for v in row.values).lower(), axis=1)
                df_rs = df_rs[mask_r]
                sts_r = [sts_r[i] for i in df_rs.index]

            st.caption(f"{len(df_rs)} relevantes · Azul=vieja · Verde=Continua · Rosa=Repetitiva")

            # Styled preview (colored like Excel)
            df_rs_reset = df_rs.reset_index(drop=True)
            def _srel(row, _s=sts_r):
                idx = row.name
                st_ = _s[idx] if idx < len(_s) else ""
                tipo = str(row.get("Tipo","")).lower() if "Tipo" in row.index else ""
                res = []
                for col in row.index:
                    if col == "Tipo":
                        res.append("background-color:#C6EFCE" if "continua" in tipo
                                   else "background-color:#FFB6C1" if "repetitiva" in tipo else "")
                    elif col == "Libranza" and st_ == "vieja":
                        res.append("background-color:#BDD7EE;font-weight:bold")
                    else:
                        res.append("")
                return res
            try:
                st.dataframe(df_rs_reset.style.apply(_srel, axis=1),
                             use_container_width=True, hide_index=True, height=400)
            except:
                st.dataframe(df_rs_reset, use_container_width=True, hide_index=True, height=400)

            # Editable section
            with st.expander("✏️ Editar relevantes"):
                df_rel_all = df_rel[disp_r].reset_index(drop=True) if disp_r else df_rel.reset_index(drop=True)
                ed_r = st.data_editor(df_rel_all, use_container_width=True,
                                      hide_index=True, num_rows="dynamic",
                                      key="prem_editor_rel", height=400)
                if st.button("💾 Guardar", key="prem_apply_rel"):
                    full_rel = st.session_state.prem_df_relevantes.copy()
                    for col in disp_r:
                        if col in ed_r.columns:
                            full_rel[col] = ed_r[col].values
                    st.session_state.prem_df_relevantes = full_rel.reset_index(drop=True)
                    st.success("✅ Guardado"); st.rerun()

    # ── TAB 4: Proyectos de Generación ───────────────────────────────
    with tabs[4]:
        st.subheader(f"Proyectos de Generación — Semana {cw}")
        df_p = st.session_state.prem_df_proyectos.copy() if st.session_state.prem_df_proyectos is not None else pd.DataFrame(columns=PROY_COLS)

        if "_fecha_fin_raw" in df_p.columns:
            df_p = df_p.drop(columns=["_fecha_fin_raw"])

        st.caption("Puede editar **Última Solicitud** y eliminar filas directamente en la tabla.")
        edited_p = st.data_editor(
            df_p.reset_index(drop=True),
            use_container_width=True, hide_index=False,
            num_rows="dynamic", key="prem_editor_proy",
            column_config={
                "Última Solicitud": st.column_config.TextColumn("Última Solicitud", width="medium"),
                "Sem Disp":  st.column_config.TextColumn("Sem Disp",  disabled=True),
                "Sem Prueba":st.column_config.TextColumn("Sem Prueba",disabled=True),
            }
        )
        if st.button("💾 Aplicar cambios", key="prem_apply_proy"):
            st.session_state.prem_df_proyectos = edited_p.iloc[edited_p.index.tolist()].reset_index(drop=True)
            st.success("Cambios aplicados"); st.rerun()



# ── Standalone execution ──────────────────────────────────────────────
try:
    st.set_page_config(
        page_title="Premisas CND", page_icon="⚡",
        layout="wide", initial_sidebar_state="expanded"
    )
except Exception:
    pass  # already set by parent app when imported
vista_premisas()
